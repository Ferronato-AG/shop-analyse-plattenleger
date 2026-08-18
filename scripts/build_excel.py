import json,csv
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter
raw={p['id']:p for p in json.load(open('data/products_raw.json'))['products']}
F=json.load(open('data/final.json'))
F.sort(key=lambda o:(o['hauptkategorie'],o['unterkategorie'],o['top_rang'] if o['top_rang']!="" else 99,-o['top_score'],o['name']))
COLS=["Hauptkategorie (neu)","Unterkategorie (neu)","Top-Rang","Relevanz 1-5","Produktname","Bild","Beschreibung (max. 500 Zeichen)","Feature / USP (Hauptbotschaft)","Sekundäre Information","Link Ferronato-Shop","Sparte/Farbe","Preis CHF exkl. MwSt","Art.-Nr.","Verfügbarkeit","Alte Shop-Unterkategorie","Begründung Top"]
def row(o):
    r=raw[o['id']]
    return [o['hauptkategorie'],o['unterkategorie'],o['top_rang'],o['top_score'],o['name'],r['image'],o['zusammenfassung'],o['usp'],o['sekundaer'],r['url'],o['sparte_farbe'],float(r['price_chf'].replace("'","")) if r['price_chf'] else "",r['artnr'],r['availability'],", ".join(r['cat_names']) or "Plattenleger (Hauptkategorie)",o['top_grund_review'] or o['top_grund']]
wb=Workbook()
def sheet(ws,rows,title):
    ws.title=title; ws.append(COLS)
    for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1F3A5F"); c.alignment=Alignment(wrap_text=True,vertical="top")
    for r in rows: ws.append(r)
    widths=[26,24,8,8,38,14,60,40,40,14,16,10,12,12,26,36]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for rw in ws.iter_rows(min_row=2):
        for c in rw: c.alignment=Alignment(wrap_text=True,vertical="top")
        for idx in (6,10):
            c=rw[idx-1]
            if c.value: c.hyperlink=c.value; c.value="Bild" if idx==6 else "Shop"; c.font=Font(color="0563C1",underline="single")
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
sheet(wb.active,[row(o) for o in F],"Alle Produkte")
sheet(wb.create_sheet(),[row(o) for o in F if o['top_rang']!=""],"Top-Produkte")
# Kategorien-Übersicht
ws=wb.create_sheet("Kategorien")
ws.append(["Hauptkategorie (neu)","Unterkategorie (neu)","Anzahl Produkte","Top-Produkte"])
import collections
g=collections.defaultdict(list)
for o in F: g[(o['hauptkategorie'],o['unterkategorie'])].append(o)
for k in sorted(g): ws.append([k[0],k[1],len(g[k]),"; ".join(o['name'] for o in g[k] if o['top_rang']!="")])
for c in ws[1]: c.font=Font(bold=True)
for i,w in enumerate([28,32,10,90],1): ws.column_dimensions[get_column_letter(i)].width=w
wb.save("data/ferronato_plattenleger_produkte.xlsx")
with open("data/ferronato_plattenleger_produkte.csv","w",newline="") as f:
    w=csv.writer(f,delimiter=";"); w.writerow(COLS); [w.writerow(row(o)) for o in F]
print("ok",len(F))
