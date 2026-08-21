import json,csv,collections
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter
raw={p['id']:p for p in json.load(open('data/all_products_raw.json'))['products']}
F=json.load(open('data/sfinal.json'))
F.sort(key=lambda o:(o['sparte'],o['hauptkategorie'],o['unterkategorie'],o['top_rang'] if o['top_rang']!="" else 99,-o['top_score'],o['name']))
COLS=["Sparte","Hauptkategorie (neu)","Unterkategorie (neu)","Top-Rang","Relevanz 1-5","Produktname","Bild","Beschreibung (max. 500 Zeichen)","Feature / USP (Hauptbotschaft)","Sekundäre Information","Link Ferronato-Shop","Preis CHF exkl. MwSt","Art.-Nr.","Verfügbarkeit","Alte Shop-Kategorien","Begründung Top"]
def row(o):
    r=raw[o['id']]
    url=f"https://shop.ferronato.ch/de/product/{o['id']}/x"
    return [o['sparte'],o['hauptkategorie'],o['unterkategorie'],o['top_rang'],o['top_score'],o['name'],r['image'],o['zusammenfassung'],o['usp'],o['sekundaer'],url,float(r['price_chf'].replace("'","")) if r['price_chf'] else "",r['artnr'],r['availability']," > ".join(r['breadcrumb']),o['top_grund_review'] or o.get('top_grund','')]
wb=Workbook()
def sheet(ws,rows,title):
    ws.title=title; ws.append(COLS)
    for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1F3A5F"); c.alignment=Alignment(wrap_text=True,vertical="top")
    for r in rows: ws.append(r)
    for i,w in enumerate([20,26,24,8,8,38,10,60,40,40,12,10,12,12,30,36],1): ws.column_dimensions[get_column_letter(i)].width=w
    for rw in ws.iter_rows(min_row=2):
        for c in rw: c.alignment=Alignment(wrap_text=True,vertical="top")
        for idx,label in ((7,"Bild"),(11,"Shop")):
            c=rw[idx-1]
            if c.value: c.hyperlink=c.value; c.value=label; c.font=Font(color="0563C1",underline="single")
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
sheet(wb.active,[row(o) for o in F],"Alle Produkte")
sheet(wb.create_sheet(),[row(o) for o in F if o['top_rang']!=""],"Top-Produkte")
ws=wb.create_sheet("Kategorien")
ws.append(["Sparte","Hauptkategorie","Unterkategorie","Anzahl","Top-Produkte"])
g=collections.defaultdict(list)
for o in F: g[(o['sparte'],o['hauptkategorie'],o['unterkategorie'])].append(o)
for k in sorted(g): ws.append([k[0],k[1],k[2],len(g[k]),"; ".join(o['name'] for o in sorted(g[k],key=lambda x:x['top_rang'] if x['top_rang']!="" else 99) if o['top_rang']!="")])
for c in ws[1]: c.font=Font(bold=True)
for i,w in enumerate([20,28,28,8,90],1): ws.column_dimensions[get_column_letter(i)].width=w
wb.save("deliverable/ferronato_gesamtshop_produkte.xlsx")
with open("deliverable/ferronato_gesamtshop_produkte.csv","w",newline="") as f:
    w=csv.writer(f,delimiter=";"); w.writerow(COLS); [w.writerow(row(o)) for o in F]
print("ok",len(F))
